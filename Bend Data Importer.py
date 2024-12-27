import os
import re
import logging
from pathlib import Path
from tkinter import Tk, filedialog
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(filename="suprafile_parser.log", level=logging.ERROR, format="%(asctime)s - %(levelname)s - %(message)s")

def parse_fixed_suprafile(file_path):
    try:
        with open(file_path, 'r') as file:
            contents = file.readlines()

        part_data = {
            "part_name": "",
            "part_length": None,
            "bends": []
        }

        parsing_bends = False
        found_pvar = False

        for line in contents:
            clean_line = line.strip()

            if clean_line.startswith("$pnum"):
                split_part_data = clean_line.split()
                try:
                    part_data["part_name"] = split_part_data[1]+split_part_data[2]
                except IndexError:
                    part_data["part_name"] = split_part_data[1]

            elif clean_line.startswith("$clra"):
                parsing_bends = True
                part_data["bends"] = []

            elif parsing_bends and clean_line and not clean_line.startswith("$"):
                bend_data = re.split(r'\s+', clean_line)
                if len(bend_data) >= 11:
                    bend_info = {
                        "degree_of_bend": float(bend_data[8])
                    }
                    part_data["bends"].append(bend_info)

            elif clean_line.startswith("$pvar"):
                found_pvar = True

            if found_pvar and clean_line.startswith("TUBL"):
                try:
                    part_data["part_length"] = float(clean_line.split()[1])
                    found_pvar = False
                except (IndexError, ValueError):
                    logging.error(f"Error extracting Part Length from {file_path}")
                    part_data["part_length"] = None

        return part_data
    except Exception as e:
        logging.error(f"Unexpected error parsing file {file_path}: {e}")
        return {
            "part_name": "",
            "part_length": None,
            "bends": []
        }


def update_spreadsheet(spreadsheet_path, suprafiles_folder, save_path):
    workbook = load_workbook(spreadsheet_path, keep_vba=True)
    worksheet = workbook['PIECEMARKS']

    for file_name in os.listdir(suprafiles_folder):
        if file_name.endswith(".$$$"):
            file_path = os.path.join(suprafiles_folder, file_name)
            parsed = parse_fixed_suprafile(file_path)
            clean_part_name = parsed["part_name"].lower().replace("+", "").strip()
            total_bend_angle = sum(bend["degree_of_bend"] for bend in parsed["bends"])

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=8, max_col=8):
                part_name_cell = row[0]
                if part_name_cell.value and part_name_cell.value.lower().replace("+", "").strip() == clean_part_name:
                    worksheet.cell(row=part_name_cell.row, column=10).value = parsed["part_length"]
                    worksheet.cell(row=part_name_cell.row, column=14).value = total_bend_angle

    workbook.save(save_path)


def main():
    Tk().withdraw()  # Hide the main Tkinter window

    file_directory = "Y:/02 job files/"
    print(Path(file_directory,"12 STARBEND FILES"))
    # Prompt for spreadsheet and suprafiles folder
    spreadsheet_path = filedialog.askopenfilename(title="Select Spreadsheet File", filetypes=[("Excel Files", "*.xlsm")],initialdir=file_directory)
    suprafiles_folder = filedialog.askdirectory(title="Select Suprafiles Folder",initialdir="Y:/12 STARBEND FILES")
    if not spreadsheet_path or not suprafiles_folder:
        print("Operation canceled.")
        return

    # Prompt for save location
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsm", filetypes=[("Excel Files", "*.xlsm")], title="Save Updated Spreadsheet",initialdir=Path(spreadsheet_path).parent)
    if not save_path:
        print("Save operation canceled.")
        return

    update_spreadsheet(spreadsheet_path, suprafiles_folder, save_path)
    print(f"Updated spreadsheet saved to {save_path}")


if __name__ == "__main__":
    main()
